"""Tests for the Excel generation module."""

import re
from io import BytesIO

import openpyxl
import pytest

from texthunter.api.schemas import MatchResult
from texthunter.core.excel import build_dataframe, generate_excel


def make_match(
    source_file="test.pdf",
    page=1,
    match_found="10\"-FG-001",
    context="...connects to 10\"-FG-001 at...",
    project_id=None,
    sheet_no=None,
) -> MatchResult:
    return MatchResult(
        source_file=source_file,
        page=page,
        match_found=match_found,
        context=context,
        project_id=project_id,
        sheet_no=sheet_no,
    )


class TestBuildDataframe:
    def test_includes_context_by_default(self):
        df = build_dataframe([make_match()])
        assert "Context (± 20 chars)" in df.columns

    def test_excludes_context_when_flag_false(self):
        df = build_dataframe([make_match()], include_context=False)
        assert "Context (± 20 chars)" not in df.columns

    def test_row_count_matches_input(self):
        matches = [make_match(page=i) for i in range(5)]
        df = build_dataframe(matches)
        assert len(df) == 5

    def test_empty_matches_returns_empty_dataframe(self):
        df = build_dataframe([])
        assert len(df) == 0
        # pd.DataFrame([]) has no columns when given an empty list; that's expected.

    def test_project_id_and_sheet_no_populated(self):
        m = make_match(project_id="2024", sheet_no="SiteA")
        df = build_dataframe([m])
        assert df.iloc[0]["Project ID"] == "2024"
        assert df.iloc[0]["Sheet No"] == "SiteA"

    def test_none_project_id_becomes_empty_string(self):
        m = make_match(project_id=None, sheet_no=None)
        df = build_dataframe([m])
        assert df.iloc[0]["Project ID"] == ""
        assert df.iloc[0]["Sheet No"] == ""


class TestGenerateExcel:
    def test_returns_bytes_io(self):
        buffer = generate_excel([make_match()])
        assert isinstance(buffer, BytesIO)

    def test_buffer_is_valid_xlsx(self):
        buffer = generate_excel([make_match()])
        wb = openpyxl.load_workbook(buffer)
        assert "Extraction Results" in wb.sheetnames

    def test_header_row_present(self):
        buffer = generate_excel([make_match()])
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Extraction Results"]
        headers = [cell.value for cell in ws[1]]
        assert "Match Found" in headers
        assert "Source File" in headers

    def test_data_row_count_matches(self):
        matches = [make_match(page=i) for i in range(10)]
        buffer = generate_excel(matches)
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Extraction Results"]
        # Row 1 is the header; rows 2..n+1 are data
        assert ws.max_row == 11

    def test_empty_matches_produces_header_only(self):
        buffer = generate_excel([])
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Extraction Results"]
        assert ws.max_row == 1  # Only the header row

    def test_column_widths_set(self):
        buffer = generate_excel([make_match()])
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Extraction Results"]
        # All columns should have an explicit width set
        for col_letter in ["A", "B", "C", "D", "E"]:
            assert ws.column_dimensions[col_letter].width > 0

    def test_column_width_capped_at_50(self):
        long_match = make_match(match_found="X" * 100)
        buffer = generate_excel([long_match])
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Extraction Results"]
        for dim in ws.column_dimensions.values():
            assert dim.width <= 50

    def test_exclude_context_column(self):
        buffer = generate_excel([make_match()], include_context=False)
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Extraction Results"]
        headers = [cell.value for cell in ws[1]]
        assert "Context (± 20 chars)" not in headers

    def test_header_row_is_frozen(self):
        buffer = generate_excel([make_match()])
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Extraction Results"]
        assert ws.freeze_panes == "A2"


class TestRegexErrorPaths:
    """Tests for error handling paths in the regex module (complementing test_regex_engine.py)."""

    def test_guess_regex_raises_value_error_for_single_example(self):
        from texthunter.core.regex import guess_regex
        with pytest.raises(ValueError, match="At least 2 examples"):
            guess_regex(["only_one"])

    def test_guess_regex_raises_value_error_for_empty_list(self):
        from texthunter.core.regex import guess_regex
        with pytest.raises(ValueError, match="At least 2 examples"):
            guess_regex([])

    def test_invalid_file_identifier_regex_raises(self):
        from texthunter.core.regex import extract_matches
        with pytest.raises(ValueError, match="Invalid file identifier regex"):
            list(extract_matches(
                text_content={"file.pdf": {1: "some text"}},
                keyword_regex=r"\d+",
                file_identifier_regex="[invalid",
            ))

    def test_extract_matches_empty_text_returns_no_matches(self):
        from texthunter.core.regex import extract_matches
        matches = list(extract_matches(
            text_content={"file.pdf": {1: ""}},
            keyword_regex=r"\d+",
        ))
        assert matches == []

    def test_extract_matches_special_chars_in_filename(self):
        from texthunter.core.regex import extract_matches
        filename = "project (2024) [rev-A].pdf"
        matches = list(extract_matches(
            text_content={filename: {1: "value 42 here"}},
            keyword_regex=r"\d+",
        ))
        assert len(matches) == 1
        assert matches[0].source_file == filename

    def test_compile_pattern_cache_is_used(self):
        """_compile_pattern should return the same compiled object for the same pattern."""
        from texthunter.core.regex import _compile_pattern
        p1 = _compile_pattern(r"\d+")
        p2 = _compile_pattern(r"\d+")
        assert p1 is p2  # lru_cache returns the same object

    def test_compile_pattern_raises_for_invalid(self):
        from texthunter.core.regex import _compile_pattern
        # Clear cache to avoid contamination from other tests
        _compile_pattern.cache_clear()
        with pytest.raises(ValueError, match="Invalid regex pattern"):
            _compile_pattern("[unclosed")
