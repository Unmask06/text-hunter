"""Excel generation utilities using Pandas and Openpyxl."""

from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from texthunter.api.schemas import MatchResult

if TYPE_CHECKING:
    from texthunter.api.schemas import SymbolDetectionResult


def build_dataframe(
    matches: list[MatchResult], include_context: bool = True
) -> pd.DataFrame:
    """Create a DataFrame from match results.

    Args:
        matches: List of MatchResult objects
        include_context: Whether to include the context column

    Returns:
        Pandas DataFrame with match data

    """
    data = []
    for match in matches:
        row = {
            "Source File": match.source_file,
            "Project ID": match.project_id or "",
            "Sheet No": match.sheet_no or "",
            "Page": match.page,
            "Match Found": match.match_found,
        }
        if include_context:
            row["Context (± 20 chars)"] = match.context
        data.append(row)

    return pd.DataFrame(data)


def generate_excel(matches: list[MatchResult], include_context: bool = True) -> BytesIO:
    """Generate an Excel file from match results.

    Args:
        matches: List of MatchResult objects
        include_context: Whether to include the context column

    Returns:
        BytesIO buffer containing the Excel file

    """
    df = build_dataframe(matches, include_context)

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extraction Results")

        # Get the workbook and worksheet for styling
        worksheet = writer.sheets["Extraction Results"]

        # Style the header row
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass

            # Cap width and add padding
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze the header row
        worksheet.freeze_panes = "A2"

    buffer.seek(0)
    return buffer


def generate_vision_excel(results: list[SymbolDetectionResult]) -> BytesIO:
    """Generate an Excel file from symbol detection results.

    Follows the same header styling as generate_excel (dark blue headers,
    freeze panes, auto-adjusted column widths).

    Args:
        results: List of SymbolDetectionResult objects from the vision pipeline.

    Returns:
        BytesIO buffer containing the Excel file.

    """
    data = [
        {
            "Symbol Type": r.symbol_name,
            "Page": r.page,
            "X": r.bbox.x,
            "Y": r.bbox.y,
            "Width": r.bbox.width,
            "Height": r.bbox.height,
            "Confidence": round(r.confidence, 4),
            "Scale": round(r.scale, 3),
            "Angle (deg)": r.angle_deg,
            "Associated Tags": ", ".join(r.associated_tags),
        }
        for r in results
    ]

    df = pd.DataFrame(data)
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Symbol Detections")
        worksheet = writer.sheets["Symbol Detections"]

        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        worksheet.freeze_panes = "A2"

    buffer.seek(0)
    return buffer
